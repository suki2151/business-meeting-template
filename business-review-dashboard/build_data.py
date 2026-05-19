"""
从 上次会议.xlsx 和 本次会议.xlsx 提取数据，生成 JSON 供 看板.html 使用。

两个文件 PPT 结构不同，需要动态检测 section 位置。
"""
import openpyxl
import json
import re
import os
import datetime as dt

BASE = r"d:\claude2"

def safe_text(cell):
    if cell is None:
        return ""
    try:
        v = cell.value
        if v is None:
            return ""
        if isinstance(v, str) and v.startswith("#"):
            return ""
        return str(v).strip()
    except:
        return ""

def safe_num(cell):
    t = safe_text(cell)
    if t == "" or t == "-":
        return 0.0
    try:
        t = t.replace(",", "").replace(" ", "")
        if t.endswith("%"):
            return float(t.replace("%", "")) / 100.0
        return float(t)
    except:
        return 0.0

def read_row_values(sheet, row, start_col, count):
    result = []
    for c in range(start_col, start_col + count):
        result.append(safe_num(sheet.cell(row=row, column=c)))
    return result


def extract_meeting_date(wb):
    """提取会议日期 - Row 2 Col 15"""
    sheet = wb["PPT（3-7页）-总体情况"]
    for r in range(1, 5):
        for c in range(14, 17):
            cell = sheet.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, dt.datetime):
                return f"{v.year}/{v.month}/{v.day}"
            if isinstance(v, str):
                m = re.match(r"(\d{4}/\d{1,2}/\d{1,2})", v)
                if m:
                    return m.group(1)
    return ""


def find_sections(sheet):
    """动态查找 PPT sheet 中各 section 的起止行。
    返回 {"P3": (start, end), "P4": (start, end), "P5": (start, end), ...}
    """
    sections = {}
    for r in range(1, min(sheet.max_row + 1, 50)):
        c1 = safe_text(sheet.cell(row=r, column=1))
        m = re.match(r'^P(\d+)', c1)
        if m:
            section = 'P' + m.group(1)
            # Find the end of data: first row after start where Col1 is empty or next P section
            end = r + 1
            while end < min(sheet.max_row + 1, 50):
                next_c1 = safe_text(sheet.cell(row=end, column=1))
                if re.match(r'^P\d+', next_c1):
                    break
                if next_c1 == '' and end > r + 5:
                    # Allow some empty rows within section
                    pass
                end += 1
            sections[section] = (r, end - 1)
    return sections


def extract_ppt_overview(wb):
    """提取 PPT（3-7页）-总体情况 — 动态检测 section 位置"""
    sheet = wb["PPT（3-7页）-总体情况"]
    sections = find_sections(sheet)

    result = {
        "p3_revenue": [],
        "p4_forecast": [],
        "p5_quarterly": [],
        "sections_found": list(sections.keys()),
    }

    def find_data_rows(section_start, min_rows=5):
        """从 section P 标记行开始，找到表头行，然后读取数据直到空行"""
        header_row = None
        for r in range(section_start + 1, section_start + 5):
            c1 = safe_text(sheet.cell(row=r, column=1))
            # 表头行通常 Col1="部门"
            if "部门" in c1:
                header_row = r
                break

        if header_row is None:
            return []

        rows = []
        for r in range(header_row + 1, header_row + 15):
            dept = safe_text(sheet.cell(row=r, column=1))
            if not dept or re.match(r'^P\d+', dept):
                break
            rows.append(r)
        return rows

    # Extract P3 (revenue data)
    if "P3" in sections:
        data_rows = find_data_rows(sections["P3"][0])
        for r in data_rows:
            dept = safe_text(sheet.cell(row=r, column=1))
            if dept:
                result["p3_revenue"].append({
                    "dept": dept,
                    "revenue_current": safe_num(sheet.cell(row=r, column=2)),
                    "revenue_target": safe_num(sheet.cell(row=r, column=3)),
                    "revenue_rate": safe_text(sheet.cell(row=r, column=4)),
                    "revenue_yoy": safe_text(sheet.cell(row=r, column=5)),
                    "self_rev_current": safe_num(sheet.cell(row=r, column=6)),
                    "self_rev_target": safe_num(sheet.cell(row=r, column=7)),
                    "self_rev_rate": safe_text(sheet.cell(row=r, column=8)),
                    "self_rev_yoy": safe_text(sheet.cell(row=r, column=9)),
                    "collection": safe_num(sheet.cell(row=r, column=10)),
                })

    # Extract P4 (forecast - only in current meeting)
    if "P4" in sections:
        data_rows = find_data_rows(sections["P4"][0])
        for r in data_rows:
            dept = safe_text(sheet.cell(row=r, column=1))
            if dept:
                result["p4_forecast"].append({
                    "dept": dept,
                    "current": safe_num(sheet.cell(row=r, column=2)),
                    "current_rate": safe_text(sheet.cell(row=r, column=3)),
                    "backlog_forecast": safe_num(sheet.cell(row=r, column=4)),
                    "backlog_rate": safe_text(sheet.cell(row=r, column=5)),
                    "pending_forecast": safe_num(sheet.cell(row=r, column=6)),
                    "pending_rate": safe_text(sheet.cell(row=r, column=7)),
                    "full_year_forecast": safe_num(sheet.cell(row=r, column=8)),
                    "full_year_rate": safe_text(sheet.cell(row=r, column=9)),
                    "prev_forecast": safe_num(sheet.cell(row=r, column=10)),
                })

    # Extract P5 (quarterly comparison)
    if "P5" in sections:
        data_rows = find_data_rows(sections["P5"][0])
        for r in data_rows:
            dept = safe_text(sheet.cell(row=r, column=1))
            if dept:
                result["p5_quarterly"].append({
                    "dept": dept,
                    "q1_initial_plan": safe_num(sheet.cell(row=r, column=2)),
                    "q1_end_plan": safe_num(sheet.cell(row=r, column=3)),
                    "q1_actual": safe_num(sheet.cell(row=r, column=4)),
                    "q1_rate": safe_text(sheet.cell(row=r, column=5)),
                    "q2_prev_plan": safe_num(sheet.cell(row=r, column=6)),
                    "q2_current_plan": safe_num(sheet.cell(row=r, column=7)),
                    "q2_actual": safe_num(sheet.cell(row=r, column=8)),
                    "q2_rate": safe_text(sheet.cell(row=r, column=9)),
                })

    return result


def extract_contract_plans(wb):
    """提取 产值及收款计划 — 列位置固定：27-38产值计划, 53-64收款计划"""
    sheet = wb["产值及收款计划"]

    contracts = []
    for r in range(5, sheet.max_row + 1):
        contract_no = safe_text(sheet.cell(row=r, column=3))
        contract_name = safe_text(sheet.cell(row=r, column=4))
        if not contract_no and not contract_name:
            continue

        contracts.append({
            "dept": safe_text(sheet.cell(row=r, column=2)),
            "contract_no": contract_no,
            "contract_name": contract_name,
            "contract_amount": safe_num(sheet.cell(row=r, column=5)),
            "dept_amount": safe_num(sheet.cell(row=r, column=6)),
            "self_amount": safe_num(sheet.cell(row=r, column=8)),
            "self_rate": safe_text(sheet.cell(row=r, column=9)),
            "tax_rate": safe_text(sheet.cell(row=r, column=10)),
            "is_leading": safe_text(sheet.cell(row=r, column=13)),
            "team": safe_text(sheet.cell(row=r, column=14)),
            "manager": safe_text(sheet.cell(row=r, column=15)),
            "contract_type": safe_text(sheet.cell(row=r, column=16)),
            "confirmed_history": safe_num(sheet.cell(row=r, column=20)),
            "confirmed_2026": safe_num(sheet.cell(row=r, column=21)),
            "remaining_value": safe_num(sheet.cell(row=r, column=22)),
            "monthly_plan": read_row_values(sheet, r, 27, 12),
            "monthly_actual": read_row_values(sheet, r, 83, 12),
            "monthly_collection_plan": read_row_values(sheet, r, 53, 12),
            "monthly_collection_actual": read_row_values(sheet, r, 109, 12),
            "collected_history": safe_num(sheet.cell(row=r, column=49)),
            "collected_2026": safe_num(sheet.cell(row=r, column=50)),
            "remaining_collection": safe_num(sheet.cell(row=r, column=51)),
        })

    return contracts


def extract_top10(wb):
    """提取 TOP10"""
    sheet = wb["PPT(12-15页）-TOP10"]
    result = {"summary": [], "top10_list": []}

    for r in range(5, 15):
        dept = safe_text(sheet.cell(row=r, column=4))
        count = int(safe_num(sheet.cell(row=r, column=5)))
        plan = safe_num(sheet.cell(row=r, column=6))
        if dept:
            result["summary"].append({"dept": dept, "contract_count": count, "plan_amount": plan})

    for r in range(5, 15):
        dept = safe_text(sheet.cell(row=r, column=8))
        name = safe_text(sheet.cell(row=r, column=9))
        amount = safe_num(sheet.cell(row=r, column=10))
        plan = safe_num(sheet.cell(row=r, column=11))
        actual = safe_num(sheet.cell(row=r, column=12))
        if name:
            result["top10_list"].append({"dept": dept, "contract_name": name, "contract_amount": amount, "plan_amount": plan, "actual": actual})

    return result


def extract_receivables_summary(wb):
    """提取应收未收"""
    sheet = wb["PPT(16-19页）-应收未收"]
    result = []
    for r in range(4, min(sheet.max_row + 1, 50)):
        dept = safe_text(sheet.cell(row=r, column=1))
        val1 = safe_num(sheet.cell(row=r, column=3))
        val2 = safe_num(sheet.cell(row=r, column=5))
        val3 = safe_num(sheet.cell(row=r, column=7))
        val4 = safe_num(sheet.cell(row=r, column=9))
        if dept:
            result.append({"dept": dept, "balance_feb11": val1, "balance_jan23": val2, "balance_2025end": val3, "balance_2024end": val4})
    return result


def extract_company_summary(wb):
    """提取 公司-汇总页 — 数据在 Col2 (部门名) 和后续列"""
    sheet = wb["公司-汇总页"]
    result_产值 = []
    result_收款 = []

    # 扫描 Col2 找 "产值" 和 "收款" 标记
    current_section = None
    for r in range(1, sheet.max_row + 1):
        c2 = safe_text(sheet.cell(row=r, column=2))
        c3 = safe_text(sheet.cell(row=r, column=3))

        if "产值" in c2:
            current_section = "产值"
            continue
        if "收款" in c2:
            current_section = "收款"
            continue

        dept = c2
        # 有效部门名检查
        valid_depts = ["数字建造中心", "智能运管中心", "数智管理中心", "福建分公司",
                       "广东分公司", "产品研发中心", "市场经营拓展中心", "合计"]
        if dept not in valid_depts:
            continue

        row_data = {
            "dept": dept,
            "annual_plan": safe_num(sheet.cell(row=r, column=3)),
            "annual_actual": safe_num(sheet.cell(row=r, column=4)),
            "months": [],
        }
        for m in range(12):
            row_data["months"].append({
                "plan": safe_num(sheet.cell(row=r, column=5 + m * 2)),
                "actual": safe_num(sheet.cell(row=r, column=6 + m * 2)),
            })

        if current_section == "产值":
            result_产值.append(row_data)
        elif current_section == "收款":
            result_收款.append(row_data)

    return {"产值": result_产值, "收款": result_收款}


def extract_contract_check(wb):
    """提取 会议合同检查"""
    if "会议合同检查" not in wb.sheetnames:
        return []
    sheet = wb["会议合同检查"]
    result = []
    for r in range(2, sheet.max_row + 1):
        cno = safe_text(sheet.cell(row=r, column=1))
        if cno:
            result.append({
                "contract_no": cno,
                "remaining": safe_num(sheet.cell(row=r, column=2)),
                "status": safe_text(sheet.cell(row=r, column=3)),
            })
    return result


def extract_all(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {
        "date": extract_meeting_date(wb),
        "ppt_overview": extract_ppt_overview(wb),
        "company_summary": extract_company_summary(wb),
        "contracts": extract_contract_plans(wb),
        "top10": extract_top10(wb),
        "receivables_summary": extract_receivables_summary(wb),
        "contract_check": extract_contract_check(wb),
    }
    wb.close()
    return data


def main():
    files = {
        "last": os.path.join(BASE, "上次会议.xlsx"),
        "current": os.path.join(BASE, "本次会议.xlsx"),
    }

    all_data = {}
    for key, path in files.items():
        print(f"Extracting {key}: ...")
        all_data[key] = extract_all(path)
        d = all_data[key]
        print(f"  Date: {d['date']}")
        print(f"  Sections: {d['ppt_overview']['sections_found']}")
        print(f"  P3: {len(d['ppt_overview']['p3_revenue'])} rows")
        print(f"  P4: {len(d['ppt_overview']['p4_forecast'])} rows")
        print(f"  P5: {len(d['ppt_overview']['p5_quarterly'])} rows")
        print(f"  Contracts: {len(d['contracts'])}")
        print(f"  TOP10: sum={len(d['top10']['summary'])} list={len(d['top10']['top10_list'])}")
        print(f"  Receivables: {len(d['receivables_summary'])}")
        print(f"  Company summary: {len(d['company_summary']['产值'])} depts")
        print()

    # Cross-reference: 本次 P4 的上期预测数 = 上次 P5 的全年预测或 P3 合计
    cur_p4 = all_data["current"]["ppt_overview"]["p4_forecast"]
    last_p3 = all_data["last"]["ppt_overview"]["p3_revenue"]
    last_p5 = all_data["last"]["ppt_overview"]["p5_quarterly"]

    # 用上次会议的 P3 合计值作为 "上期预测" 的参考
    dept_map = {}
    for r in last_p3:
        dept_map[r["dept"]] = r
    for r in last_p5:
        if r["dept"] not in dept_map:
            dept_map[r["dept"]] = r

    for row in cur_p4:
        dept = row["dept"]
        if dept in dept_map:
            # 使用上次会议的营收完成值作为上期参考
            row["prev_forecast"] = dept_map[dept].get("revenue_current", 0)
        elif dept == "合计":
            # 合计使用上次 P3 合计
            total_row = [r for r in last_p3 if r["dept"] == "合计"]
            if total_row:
                row["prev_forecast"] = total_row[0]["revenue_current"]

    output_path = os.path.join(BASE, "data.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"Data -> {output_path}, {os.path.getsize(output_path)/1024:.1f} KB")


if __name__ == "__main__":
    main()
